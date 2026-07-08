#!/usr/bin/env python3
"""Build the Supportability ROI model as an .xlsx workbook.

Two ledgers — cost saving and revenue impact — each with transparent formulas,
then a summary. Requires openpyxl (`pip install openpyxl`).

Usage:
    python build_roi_model.py
    -> writes supportability_roi_model.xlsx next to this script
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---- Inputs (mirror inputs.md) ------------------------------------------------
INPUTS = {
    "Annual case volume": 10000,
    "Cost per case (AUD)": 85,
    "Escalation rate baseline": 0.20,
    "Escalation cost premium (AUD)": 160,
    "Mean handle time (min)": 55,
    "Loaded hourly rate (AUD)": 75,
    "Supported ARR (AUD)": 28_000_000,
    "Renewal rate baseline": 0.89,
    "Milestone accounts / yr": 600,
    "Avg expansion value (AUD)": 5000,
    "Expansion rate baseline": 0.06,
    "Deflection target": 0.18,
    "Escalation reduction": 0.30,
    "Handle-time reduction": 0.15,
    "Renewal lift (pts)": 0.015,
    "Expansion enablement (pts)": 0.02,
}

HEADER = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5496")
MONEY = PatternFill("solid", fgColor="E2EFDA")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)


def _style_header(cell, fill=HEADER):
    cell.fill = fill
    cell.font = WHITE
    cell.alignment = Alignment(horizontal="left", vertical="center")


def build() -> Path:
    i = INPUTS
    wb = Workbook()

    # ---- Inputs sheet ----
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Supportability ROI Model — Inputs"
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([])
    ws.append(["Input", "Value"])
    for c in ws[3]:
        _style_header(c)
    for k, v in i.items():
        ws.append([k, v])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16

    # ---- Cost ledger ----
    cs = wb.create_sheet("Cost Saving")
    deflected = i["Annual case volume"] * i["Deflection target"]
    deflect_saving = deflected * i["Cost per case (AUD)"]
    esc_reduced = (
        i["Annual case volume"] * i["Escalation rate baseline"] * i["Escalation reduction"]
    )
    esc_saving = esc_reduced * i["Escalation cost premium (AUD)"]
    ht_hours_saved = (
        i["Annual case volume"]
        * (i["Mean handle time (min)"] / 60)
        * i["Handle-time reduction"]
    )
    ht_saving = ht_hours_saved * i["Loaded hourly rate (AUD)"]
    cost_total = deflect_saving + esc_saving + ht_saving

    cs.append(["Cost-Saving Ledger"])
    cs["A1"].font = Font(size=14, bold=True)
    cs.append([])
    cs.append(["Lever", "Mechanism", "Annual saving (AUD)"])
    for c in cs[3]:
        _style_header(c)
    cs.append(["Case deflection", f"{deflected:,.0f} cases never opened", deflect_saving])
    cs.append(["Escalation reduction", f"{esc_reduced:,.0f} fewer escalations", esc_saving])
    cs.append(["Faster resolution", f"{ht_hours_saved:,.0f} labour hours saved", ht_saving])
    cs.append(["TOTAL COST SAVING", "", cost_total])
    for col in ("A", "B", "C"):
        cs[f"{col}7"].font = BOLD
    cs["C7"].fill = MONEY
    cs.column_dimensions["A"].width = 22
    cs.column_dimensions["B"].width = 34
    cs.column_dimensions["C"].width = 20

    # ---- Revenue ledger ----
    rv = wb.create_sheet("Revenue Impact")
    protected = i["Supported ARR (AUD)"] * i["Renewal lift (pts)"]
    enabled = (
        i["Milestone accounts / yr"]
        * i["Expansion enablement (pts)"]
        * i["Avg expansion value (AUD)"]
    )
    rev_total = protected + enabled

    rv.append(["Revenue-Impact Ledger"])
    rv["A1"].font = Font(size=14, bold=True)
    rv.append([])
    rv.append(["Lever", "Mechanism", "Annual impact (AUD)"])
    for c in rv[3]:
        _style_header(c, SUB)
    rv.append(
        ["Retention lift", f"+{i['Renewal lift (pts)']*100:.1f}pt renewal on supported ARR", protected]
    )
    rv.append(
        ["Expansion enabled", f"+{i['Expansion enablement (pts)']*100:.0f}pt expansion on milestone accts", enabled]
    )
    rv.append(["TOTAL REVENUE IMPACT", "", rev_total])
    for col in ("A", "B", "C"):
        rv[f"{col}6"].font = BOLD
    rv["C6"].fill = MONEY
    rv.column_dimensions["A"].width = 22
    rv.column_dimensions["B"].width = 42
    rv.column_dimensions["C"].width = 20

    # ---- Summary ----
    sm = wb.create_sheet("Summary")
    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))
    sm.append(["Supportability Value — Summary"])
    sm["A1"].font = Font(size=16, bold=True)
    sm.append([])
    sm.append(["Ledger", "Annual value (AUD)"])
    for c in sm[3]:
        _style_header(c)
    sm.append(["Cost saving", cost_total])
    sm.append(["Revenue impact", rev_total])
    sm.append(["TOTAL VALUE", cost_total + rev_total])
    sm["A6"].font = Font(bold=True, size=12)
    sm["B6"].font = Font(bold=True, size=12)
    sm["B6"].fill = MONEY
    for r in range(4, 7):
        sm[f"B{r}"].number_format = "#,##0"
    sm.column_dimensions["A"].width = 22
    sm.column_dimensions["B"].width = 22

    # money formats
    for sheet, col, rng in (("Cost Saving", "C", range(4, 8)), ("Revenue Impact", "C", range(4, 7))):
        for r in rng:
            wb[sheet][f"{col}{r}"].number_format = "#,##0"

    out = Path(__file__).parent / "supportability_roi_model.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    p = build()
    print(f"Wrote {p}")
